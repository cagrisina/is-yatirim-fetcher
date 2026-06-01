# Excel sayfa yazımı: Config, Raw_data (uzun format + anlık metrikler), Sinyal (Python ile hesaplanan değerler).

import statistics

from openpyxl.styles import PatternFill

_NO_FILL = PatternFill(fill_type=None)

RAW_DATA_SHEET = "Raw_data"

# (Raw_data "Metrik" kodu, hist hücredeki indeks: pddd, fd_favok, borc_ozk, ihracat_orani)
METRICS = [
    ("PDDD", 0),
    ("FD_FAVOK", 1),
    ("BORC_OZK", 2),
    ("IHRACAT_ORANI", 3),
]

MANAGED_OUTPUT_SHEETS = ("Raw_data", "Yabancı Oranı", "Sinyal")

# Eski kitaplarda kalan Config F dönem etiketleri — artık kullanılmıyor; her çalıştırmada temizlenir.
_CONFIG_LEGACY_PERIOD_COL = 6
_CONFIG_LEGACY_PERIOD_MAX_ROW = 50


def ensure_config_sheet(wb):
    if "Config" in wb.sheetnames:
        _clear_config_legacy_period_column(wb["Config"])
        return
    ws = wb.create_sheet("Config")
    ws["A1"], ws["B1"] = "analiz_donemi", 12
    ws["A2"], ws["B2"] = "ucuz_esik", 25
    ws["A3"], ws["B3"] = "pahali_esik", 75


def _clear_config_legacy_period_column(ws):
    col = _CONFIG_LEGACY_PERIOD_COL
    for r in range(1, _CONFIG_LEGACY_PERIOD_MAX_ROW + 1):
        ws.cell(row=r, column=col, value=None)


def read_config_signal_params(wb):
    """Config!B1:B3 — analiz_donemi (çekilecek çeyrek sayısı + Sinyal penceresi / pct paydası), ucuz_esik, pahali_esik."""
    ensure_config_sheet(wb)
    ws = wb["Config"]

    def _num(cell, default, as_int=False):
        v = ws[cell].value
        if v is None or v == "":
            return default
        try:
            x = float(v)
            return int(x) if as_int else x
        except (TypeError, ValueError):
            return default

    return (
        _num("B1", 12, as_int=True),
        _num("B2", 25.0),
        _num("B3", 75.0),
    )


def remove_facts_sheet_if_present(wb):
    if "Facts" in wb.sheetnames:
        wb.remove(wb["Facts"])


def clear_managed_output_sheets_body(wb, max_row=5000, max_col=30):
    """Main hariç kodun doldurduğu sayfalarda satır 2+ değer ve dolgu temizliği (başlık satırı korunur)."""
    for name in MANAGED_OUTPUT_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)


def _clear_matrix_sheet(ws, max_row=None, max_col=None):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)


def write_raw_data(wb, period_labels, matrix_rows, snapshot_rows):
    """
    Raw_data: Ticker | Metrik | Dönem | Değer (uzun format).

    Önce çeyreklik geçmiş metrikler (PDDD, FD_FAVOK, BORC_OZK), sonra anlık snapshot satırları.
    snapshot_rows: (ticker, metrik, dönem, değer, number_format_str|None) listesi.
    """
    n = len(period_labels)
    n_metrics = len(METRICS)
    snap_n = len(snapshot_rows)
    data_rows = max(1, len(matrix_rows) * max(n, 1) * n_metrics) + snap_n + 5

    if RAW_DATA_SHEET not in wb.sheetnames:
        wb.create_sheet(RAW_DATA_SHEET)
    ws = wb[RAW_DATA_SHEET]

    _clear_matrix_sheet(ws, max(data_rows + 50, 500), 10)

    ws.cell(row=1, column=1, value="Ticker")
    ws.cell(row=1, column=2, value="Metrik")
    ws.cell(row=1, column=3, value="Dönem")
    ws.cell(row=1, column=4, value="Değer")

    out_r = 2

    for _row_idx, symbol, cells in matrix_rows:
        assert len(cells) == len(period_labels), (
            f"Raw_data hizası: {symbol!r} cells={len(cells)} period_labels={len(period_labels)}"
        )
        for j in range(n):
            period = period_labels[j] if j < len(period_labels) else ""
            trip = cells[j] if j < len(cells) else (None, None, None, None)
            for metric_code, idx in METRICS:
                v = trip[idx] if idx < len(trip) else None
                if metric_code == "IHRACAT_ORANI":
                    if period.endswith("/12"):
                        display_period = period.split("/")[0]
                    else:
                        continue
                else:
                    display_period = period

                ws.cell(row=out_r, column=1, value=symbol)
                ws.cell(row=out_r, column=2, value=metric_code)
                ws.cell(row=out_r, column=3, value=display_period)
                c = ws.cell(row=out_r, column=4, value=v)
                if isinstance(v, (int, float)):
                    if metric_code == "IHRACAT_ORANI":
                        c.number_format = "0.00%"
                    else:
                        c.number_format = "0.00"
                out_r += 1

    for ticker, metrik, donem, deger, nf in snapshot_rows:
        ws.cell(row=out_r, column=1, value=ticker)
        ws.cell(row=out_r, column=2, value=metrik)
        ws.cell(row=out_r, column=3, value=donem)
        c = ws.cell(row=out_r, column=4, value=deger)
        if nf and isinstance(deger, (int, float)):
            c.number_format = nf
        elif isinstance(deger, (int, float)):
            c.number_format = "0.00"
        out_r += 1


def _metric_window_floats(cells, metric_idx, window_n):
    out = []
    for j in range(min(window_n, len(cells))):
        trip = cells[j]
        v = trip[metric_idx] if metric_idx < len(trip) else None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            out.append(float(v))
    return out


def _current_metric(cells, metric_idx):
    if not cells:
        return None
    trip = cells[0]
    v = trip[metric_idx] if metric_idx < len(trip) else None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _median_or_none(vals):
    if not vals:
        return None
    return float(statistics.median(vals))


def _excel_style_pct_in_window(window_vals, cur, denom_b1):
    """Eski Excel: SUM(--(v<=cur))/B1; payda her zaman analiz_donemi (B1)."""
    if denom_b1 <= 0 or cur is None:
        return None
    cnt = sum(1 for v in window_vals if v <= cur)
    return cnt / float(denom_b1)


def _signal_ucuz_pahali_neutral(pct, ucuz_esik, pahali_esik):
    if pct is None:
        return ""
    lo = ucuz_esik / 100.0
    hi = pahali_esik / 100.0
    if pct < lo:
        return "UCUZ"
    if pct > hi:
        return "PAHALI"
    return "NÖTR"


def _genel_signal(score, ucuz_esik, pahali_esik):
    if score is None:
        return ""
    if score > pahali_esik:
        return "UCUZ"
    if score < ucuz_esik:
        return "PAHALI"
    return "NÖTR"


def write_sinyal_sheet(wb, tr_tasks, tr_fin_results_map, matrix_rows, period_labels):
    """
    Sinyal: Raw_data ile aynı kaynaktan (matrix_rows) Python ile hesaplanır; Excel 365 gerekmez.
    A sütunu Main!A ile senkron; B (Sektör) yfinance değeri olarak yazılır.
    """
    window_cfg, ucuz_esik, pahali_esik = read_config_signal_params(wb)
    n_labels = len(period_labels)
    window_n = min(int(window_cfg), n_labels) if n_labels else 0

    sym_to_cells = {sym: cells for _r, sym, cells in matrix_rows}

    name = "Sinyal"
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    ws = wb[name]

    headers = {
        1: "Ticker",
        2: "Sektör",
        3: "Banka mı?",
        4: "Güncel PD/DD",
        5: "Medyan PD/DD",
        6: "PD/DD %ile",
        7: "PD/DD Sinyal",
        8: "Güncel FD/FAVÖK",
        9: "Medyan FD/FAVÖK",
        10: "FD/FAVÖK %ile",
        11: "FD/FAVÖK Sinyal",
        12: "Güncel Borç/Özsermaye",
        13: "Medyan Borç/Özsermaye",
        14: "Borç/Özser. %ile",
        15: "Borç/Özser. Sinyal",
        16: "Skor",
        17: "Genel Sinyal",
    }
    max_r = 2
    for r, _ in tr_tasks:
        max_r = max(max_r, r)
    _clear_matrix_sheet(ws, max(max_r + 2, 50), 20)

    for col, text in headers.items():
        ws.cell(row=1, column=col, value=text)

    for row_idx, sym in tr_tasks:
        r = row_idx
        fin = tr_fin_results_map.get(sym) or {}
        data = fin.get("data") or {}
        is_bank = bool(data.get("is_bank", False))
        c_val = "EVET" if is_bank else "HAYIR"

        ws.cell(row=r, column=1, value=sym)
        ws.cell(row=r, column=2, value=data.get("industry", ""))

        ws.cell(row=r, column=3, value=c_val)

        cells = sym_to_cells.get(sym) or []

        # --- PD/DD (0) ---
        w_p = _metric_window_floats(cells, 0, window_n)
        cur_p = _current_metric(cells, 0)
        med_p = _median_or_none(w_p)
        pct_p = _excel_style_pct_in_window(w_p, cur_p, window_cfg)
        sig_p = _signal_ucuz_pahali_neutral(pct_p, ucuz_esik, pahali_esik)

        ws.cell(row=r, column=4, value=cur_p if cur_p is not None else "")
        ws.cell(row=r, column=5, value=med_p if med_p is not None else "")
        ws.cell(row=r, column=6, value=pct_p if pct_p is not None else "")
        ws.cell(row=r, column=7, value=sig_p)

        for c in (4, 5, 6):
            x = ws.cell(row=r, column=c).value
            if isinstance(x, (int, float)):
                ws.cell(row=r, column=c).number_format = "0.00"

        # --- FD/FAVÖK (1), Borç (2) — bankada boş ---
        if is_bank:
            for c in range(8, 16):
                ws.cell(row=r, column=c, value="")
            score = None
            if pct_p is not None:
                score = round((1.0 - pct_p) * 100.0)
        else:
            w_f = _metric_window_floats(cells, 1, window_n)
            cur_f = _current_metric(cells, 1)
            med_f = _median_or_none(w_f)
            pct_f = _excel_style_pct_in_window(w_f, cur_f, window_cfg)
            sig_f = _signal_ucuz_pahali_neutral(pct_f, ucuz_esik, pahali_esik)

            ws.cell(row=r, column=8, value=cur_f if cur_f is not None else "")
            ws.cell(row=r, column=9, value=med_f if med_f is not None else "")
            ws.cell(row=r, column=10, value=pct_f if pct_f is not None else "")
            ws.cell(row=r, column=11, value=sig_f)

            w_b = _metric_window_floats(cells, 2, window_n)
            cur_b = _current_metric(cells, 2)
            med_b = _median_or_none(w_b)
            pct_b = _excel_style_pct_in_window(w_b, cur_b, window_cfg)
            sig_b = _signal_ucuz_pahali_neutral(pct_b, ucuz_esik, pahali_esik)

            ws.cell(row=r, column=12, value=cur_b if cur_b is not None else "")
            ws.cell(row=r, column=13, value=med_b if med_b is not None else "")
            ws.cell(row=r, column=14, value=pct_b if pct_b is not None else "")
            ws.cell(row=r, column=15, value=sig_b)

            for c in (8, 9, 10, 12, 13, 14):
                x = ws.cell(row=r, column=c).value
                if isinstance(x, (int, float)):
                    ws.cell(row=r, column=c).number_format = "0.00"

            score = None
            if pct_p is not None and pct_f is not None and pct_b is not None:
                score = round(
                    ((1.0 - pct_p) * 0.4 + (1.0 - pct_f) * 0.4 + (1.0 - pct_b) * 0.2) * 100.0
                )

        ws.cell(row=r, column=16, value=score if score is not None else "")
        ws.cell(row=r, column=17, value=_genel_signal(score, ucuz_esik, pahali_esik))

        if isinstance(ws.cell(row=r, column=16).value, (int, float)):
            ws.cell(row=r, column=16).number_format = "0"
