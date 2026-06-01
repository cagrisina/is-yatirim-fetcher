import os
import re
import sys
import time
import requests
from requests.adapters import HTTPAdapter
import pandas as pd
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook
import concurrent.futures
import datetime
import yfinance as yf

import excel_sheets

# ==============================================================================
# 0. PAYLAŞIMLI HTTP SESSION (BÜYÜK CONNECTION POOL)
# ==============================================================================
# Varsayılan requests.get() host başına sadece 10 bağlantı açar.
# 50 thread ile paralel çalışabilmek için pool boyutunu büyütüyoruz.
_session = requests.Session()
_adapter = HTTPAdapter(pool_connections=200, pool_maxsize=200, max_retries=1)
_session.mount("https://", _adapter)
_session.mount("http://", _adapter)

# ==============================================================================
# 1. ORTAK YARDIMCI FONKSİYONLAR (TARİH & DÖNEM)
# ==============================================================================

def build_ttm_periods(y, m):
    if m == 12: return [(y, 12), (y, 9), (y, 6), (y, 3)], f"{y}/12", False 
    else:
        filler_m = m - 3 if m > 3 else 12
        filler_y = y if m > 3 else y - 1
        return [(y, m), (y-1, 12), (y-1, m), (filler_y, filler_m)], f"{y}/{m:02d}", True

def get_isyatirim_periods(period_str):
    p_str = str(period_str).strip().upper()
    if p_str.isdigit() and len(p_str) == 4: return build_ttm_periods(int(p_str), 12)
    if "Q" in p_str:
        try:
            parts = p_str.split("Q")
            y, m = int(re.sub(r'\D', '', parts[0])), int(re.sub(r'\D', '', parts[1])) * 3
            return build_ttm_periods(y, m)
        except: pass
    if "/" in p_str:
        try:
            parts = p_str.split("/")
            y, m = int(re.sub(r'\D', '', parts[0])), int(re.sub(r'\D', '', parts[1].split()[0]))
            if m in [3, 6, 9, 12]: return build_ttm_periods(y, m)
        except: pass
    return None, None, False


def get_hist_price(t_sym, year, month):
    try:
        next_month = month + 1 if month < 12 else 1
        next_year = year if month < 12 else year + 1
        data = yf.Ticker(f"{t_sym}.IS").history(
            start=f"{year}-{month:02d}-01",
            end=f"{next_year}-{next_month:02d}-01",
        )
        if not data.empty:
            return float(data["Close"].iloc[-1])
    except Exception:
        pass
    return ""


def parse_resolved_quarter(resolved_p):
    if not resolved_p or resolved_p == "Bulunamadı":
        return None
    try:
        parts = str(resolved_p).split("/")[0:2]
        y = int(re.sub(r"\D", "", parts[0]))
        q = int(re.sub(r"\D", "", parts[1].split()[0]))
        if q in (3, 6, 9, 12):
            return y, q
    except Exception:
        pass
    return None


def build_quarter_targets(y, q, n):
    out = []
    cy, cq = y, q
    for _ in range(n):
        out.append((cy, cq))
        cq -= 3
        if cq <= 0:
            cq = 12
            cy -= 1
    return out


def resolve_quarter_list_for_workbook(tr_tasks, n):
    for _r, sym in tr_tasks:
        df_raw, resolved_p, _partial, _group = fetch_tr_raw_data(sym)
        if df_raw is None:
            continue
        pq = parse_resolved_quarter(resolved_p)
        if pq:
            return build_quarter_targets(pq[0], pq[1], n)
    now = datetime.datetime.now()
    y, m = now.year, now.month
    q = 9 if m >= 11 else 6 if m >= 8 else 3 if m >= 5 else 12
    if q == 12:
        y -= 1
    return build_quarter_targets(y, q, n)


def period_label(y, q):
    return f"{y}/{q:02d}"


# ==============================================================================
# 2. TÜRKİYE (TR) MOTORU (SAF SAYILARLA)
# ==============================================================================

def fetch_tr_raw_data(symbol, period_val=None):
    clean_symbol = symbol.split('.')[0].strip().upper()
    url = "https://www.isyatirim.com.tr/_layouts/15/IsYatirim.Website/Common/Data.aspx/MaliTablo"
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json", "X-Requested-With": "XMLHttpRequest"}
    fin_groups = ["XI_29", "UFRS_K", "UFRS"] 

    if period_val is not None:
        y, q = period_val
        attempts = 1
    else:
        now = datetime.datetime.now()
        y, m = now.year, now.month
        q = 9 if m >= 11 else 6 if m >= 8 else 3 if m >= 5 else 12
        if q == 12: y -= 1
        attempts = 4
    
    for attempt in range(attempts):
        periods_list, resolved_name, is_partial = build_ttm_periods(y, q)
        for group in fin_groups:
            params = {"companyCode": clean_symbol, "exchange": "TRY", "financialGroup": group}
            for i, (py, pm) in enumerate(periods_list, start=1):
                params[f"year{i}"] = str(py); params[f"period{i}"] = str(pm)
            try:
                res = _session.get(url, params=params, headers=headers, timeout=10)
                if res.status_code == 200:
                    data = res.json()
                    if "value" in data and data["value"]:
                        df = pd.DataFrame(data["value"])
                        if not df[df['itemDescTr'].str.contains("Özkaynak|Sermaye", case=False, na=False)].empty:
                            ret_name = resolved_name if period_val is not None else f"{resolved_name} (En Güncel)"
                            return df, ret_name, is_partial, group
            except: pass
        if period_val is None:
            q -= 3
            if q <= 0: q = 12; y -= 1
    return None, "Bulunamadı", False, None

def extract_tr_value(df, keyword_contains, regex=False):
    v1, v2, v3, v4 = 0.0, 0.0, 0.0, 0.0
    mask = df['itemDescTr'].fillna('').str.contains(keyword_contains, case=False, na=False, regex=regex)
    if not df[mask].empty:
        row = df[mask].iloc[0]
        try: v1 = float(row.get('value1') or 0)
        except: pass
        try: v2 = float(row.get('value2') or 0)
        except: pass
        try: v3 = float(row.get('value3') or 0)
        except: pass
        try: v4 = float(row.get('value4') or 0)
        except: pass
    return v1, v2, v3, v4

def parse_tr_data(df, resolved_period_name, is_partial, group):
    if df is None or df.empty: return None
    is_bank = (group in ["UFRS", "UFRS_K"])

    try: period_num = int(resolved_period_name.split('/')[1].split()[0])
    except: period_num = 12

    def calc_ttm_and_q(v1, v2, v3, v4):
        ttm = (v1 + v2 - v3) if is_partial else v1
        if period_num == 12: q = v1 - v2
        elif period_num == 3: q = v1
        else: q = v1 - v4
        return ttm, q

    def get_val_by_code(code_str):
        v1, v2, v3, v4 = 0.0, 0.0, 0.0, 0.0
        if 'itemCode' not in df.columns: return v1, v2, v3, v4
        mask = df['itemCode'].fillna('').astype(str).str.strip().str.upper() == code_str
        if not df[mask].empty:
            row = df[mask].iloc[0]
            try: v1 = float(row.get('value1') or 0)
            except: pass
            try: v2 = float(row.get('value2') or 0)
            except: pass
            try: v3 = float(row.get('value3') or 0)
            except: pass
            try: v4 = float(row.get('value4') or 0)
            except: pass
        return v1, v2, v3, v4

    item_2oa_v1 = 0.0
    if 'itemCode' in df.columns:
        mask_2oa = df['itemCode'].fillna('').str.upper() == '2OA'
        if not df[mask_2oa].empty:
            try: item_2oa_v1 = float(df[mask_2oa].iloc[0].get('value1') or 0)
            except: pass

    ozk_v1, _, _, _ = extract_tr_value(df, "Ana Ortaklığa Ait Özkaynaklar")
    if ozk_v1 == 0: ozk_v1, _, _, _ = extract_tr_value(df, "Özkaynaklar")

    if is_bank:
        net_op_inc_v1, net_op_inc_v2, net_op_inc_v3, net_op_inc_v4 = get_val_by_code("3H")
        if net_op_inc_v1 == 0 and net_op_inc_v2 == 0 and net_op_inc_v3 == 0:
            net_op_inc_v1, net_op_inc_v2, net_op_inc_v3, net_op_inc_v4 = extract_tr_value(df, "NET FAALİYET KARI/ZARARI")
        net_op_inc_ttm, net_op_inc_q = calc_ttm_and_q(net_op_inc_v1, net_op_inc_v2, net_op_inc_v3, net_op_inc_v4)

        ebit_v1, ebit_v2, ebit_v3, ebit_v4 = extract_tr_value(df, "NET FAALİYET KARI/ZARARI")
        ebit_ttm, ebit_q = calc_ttm_and_q(ebit_v1, ebit_v2, ebit_v3, ebit_v4)
        return {
            "debt_equity": "", 
            "ebit_ttm": ebit_ttm,
            "ebit_q": ebit_q,
            "net_op_income_ttm": net_op_inc_ttm,
            "net_op_income_q": net_op_inc_q,
            "op_cash_flow_ttm": "", 
            "op_cash_flow_q": "", 
            "net_fx": "", 
            "export_ratio_ttm": "", 
            "export_ratio_q": "", 
            "item_2oa": item_2oa_v1,
            "ozk": ozk_v1,
            "net_borc": "",
            "favok_ttm": "",
            "favok_q": "",
            "resolved_period": resolved_period_name, "is_bank": True
        }
    else:
        kv_borc_v1, _, _, _ = get_val_by_code("2AA")
        uv_borc_v1, _, _, _ = get_val_by_code("2BA")
        borc_v1 = kv_borc_v1 + uv_borc_v1
        
        nakit_v1, _, _, _ = get_val_by_code("1AA")
        fin_yat_kisa_v1, _, _, _ = get_val_by_code("1AB")
        fin_yat_uzun_v1, _, _, _ = get_val_by_code("1BC")
        
        net_borc = borc_v1 - (nakit_v1 + fin_yat_kisa_v1 + fin_yat_uzun_v1)
        
        ebit_v1, ebit_v2, ebit_v3, ebit_v4 = get_val_by_code("3DF")
        if ebit_v1 == 0 and ebit_v2 == 0 and ebit_v3 == 0:
            ebit_v1, ebit_v2, ebit_v3, ebit_v4 = extract_tr_value(df, r"^\s*FAALİYET KARI\s*(?:\(ZARARI\))?$", regex=True)
        ebit_ttm, ebit_q = calc_ttm_and_q(ebit_v1, ebit_v2, ebit_v3, ebit_v4)
        
        net_op_inc_v1, net_op_inc_v2, net_op_inc_v3, net_op_inc_v4 = get_val_by_code("3H")
        if net_op_inc_v1 == 0 and net_op_inc_v2 == 0 and net_op_inc_v3 == 0:
            net_op_inc_v1, net_op_inc_v2, net_op_inc_v3, net_op_inc_v4 = extract_tr_value(df, r"^\s*NET FAALİYET KARI\s*(?:\(ZARARI\))?$", regex=True)
        net_op_inc_ttm, net_op_inc_q = calc_ttm_and_q(net_op_inc_v1, net_op_inc_v2, net_op_inc_v3, net_op_inc_v4)

        amort_v1, amort_v2, amort_v3, amort_v4 = get_val_by_code("4B")
        if amort_v1 == 0 and amort_v2 == 0 and amort_v3 == 0:
            amort_v1, amort_v2, amort_v3, amort_v4 = get_val_by_code("4CAB")
        amort_ttm, amort_q = calc_ttm_and_q(amort_v1, amort_v2, amort_v3, amort_v4)
        
        favok_ttm = ebit_ttm + amort_ttm
        favok_q = ebit_q + amort_q
        
        net_fx, _, _, _ = extract_tr_value(df, "Net Yabancı Para Pozisyonu")
        
        cf_v1, cf_v2, cf_v3, cf_v4 = extract_tr_value(df, "İşletme Faaliyetlerinden Kaynaklanan")
        cf_ttm, cf_q = calc_ttm_and_q(cf_v1, cf_v2, cf_v3, cf_v4)
        
        ihracat_v1, ihracat_v2, ihracat_v3, ihracat_v4 = get_val_by_code("4BD")
        if ihracat_v1 == 0 and ihracat_v2 == 0 and ihracat_v3 == 0:
            ihracat_v1, ihracat_v2, ihracat_v3, ihracat_v4 = extract_tr_value(df, "Yurtdışı Satışlar")
        ihracat_ttm, ihracat_q = calc_ttm_and_q(ihracat_v1, ihracat_v2, ihracat_v3, ihracat_v4)
        
        hasilat_v1, hasilat_v2, hasilat_v3, hasilat_v4 = get_val_by_code("3C")
        if hasilat_v1 == 0 and hasilat_v2 == 0 and hasilat_v3 == 0:
            hasilat_v1, hasilat_v2, hasilat_v3, hasilat_v4 = extract_tr_value(df, "Satış Gelirleri")
        hasilat_ttm, hasilat_q = calc_ttm_and_q(hasilat_v1, hasilat_v2, hasilat_v3, hasilat_v4)
        
        export_ratio_ttm = (ihracat_ttm / hasilat_ttm) if hasilat_ttm != 0 else 0.0
        export_ratio_q = (ihracat_q / hasilat_q) if hasilat_q != 0 else 0.0

        return {
            "debt_equity": (borc_v1 / ozk_v1) if ozk_v1 != 0 else "N/A",
            "ebit_ttm": ebit_ttm, 
            "ebit_q": ebit_q, 
            "net_op_income_ttm": net_op_inc_ttm,
            "net_op_income_q": net_op_inc_q,
            "op_cash_flow_ttm": cf_ttm, 
            "op_cash_flow_q": cf_q, 
            "net_fx": net_fx,
            "export_ratio_ttm": export_ratio_ttm,
            "export_ratio_q": export_ratio_q,
            "item_2oa": item_2oa_v1,
            "ozk": ozk_v1,
            "net_borc": net_borc,
            "favok_ttm": favok_ttm,
            "favok_q": favok_q,
            "resolved_period": resolved_period_name, "is_bank": False
        }



# ==============================================================================
# 4. YABANCI TAKAS ORANI MOTORU
# ==============================================================================

def get_yabanci_oran_dates(test_symbol):
    today = pd.Timestamp.today().normalize()
    
    valid_end = None
    for i in range(6):
        test_end = today - pd.Timedelta(days=i)
        end_str = test_end.strftime("%d-%m-%Y")
        res = fetch_yabanci_oran(test_symbol, end_str, end_str)
        if res.get("end", "") != "":
            valid_end = test_end
            break
            
    if valid_end is None:
        valid_end = today - BDay(1)
        
    valid_start = None
    base_start = today - pd.Timedelta(days=30)
    for i in range(6):
        test_start = base_start + pd.Timedelta(days=i)
        start_str = test_start.strftime("%d-%m-%Y")
        res = fetch_yabanci_oran(test_symbol, start_str, start_str)
        if res.get("end", "") != "":
            valid_start = test_start
            break
            
    if valid_start is None:
        valid_start = base_start
        
    return valid_start.strftime("%d-%m-%Y"), valid_end.strftime("%d-%m-%Y")

def fetch_yabanci_oran(symbol, start_date, end_date):
    clean_symbol = symbol.split('.')[0].strip().upper()
    url = "https://www.isyatirim.com.tr/_layouts/15/IsYatirim.Website/StockInfo/CompanyInfoAjax.aspx/GetYabanciOranlarXHR"
    payload = {"baslangicTarih": start_date, "bitisTarihi": end_date, "sektor": None, "endeks": "09", "hisse": clean_symbol}
    headers = {"accept": "application/json", "content-type": "application/json; charset=UTF-8", "x-requested-with": "XMLHttpRequest"}

    try:
        res = _session.post(url, json=payload, headers=headers, timeout=10)
        if res.status_code == 200 and "d" in res.json() and len(res.json()["d"]) > 0:
            item = res.json()["d"][0]
            # Değerleri % formatında Excel'de gösterebilmek için 100'e bölüyoruz
            return {"start": item.get("YAB_ORAN_START", 0)/100, "end": item.get("YAB_ORAN_END", 0)/100, "change": item.get("DEGISIM", 0)/100, "effect": item.get("ETKI", 0)/100, "price": item.get("PRICE_TL", 0)}
    except: pass
    return {"start": "", "end": "", "change": "", "effect": "", "price": ""}


# ==============================================================================
# 5. MULTITHREAD GÖREV YÖNETİCİLERİ
# ==============================================================================

def task_tr_stock(row_idx, symbol):
    df_raw, resolved_p, is_partial, group = fetch_tr_raw_data(symbol) # period_val ignored
    result = parse_tr_data(df_raw, resolved_p, is_partial, group)
    if result is None:
        result = {k: "" for k in ['debt_equity', 'ebit_ttm', 'ebit_q', 'op_cash_flow_ttm', 'op_cash_flow_q', 'net_fx', 'export_ratio_ttm', 'export_ratio_q', 'resolved_period', 'item_2oa', 'ozk', 'net_borc', 'favok_ttm', 'favok_q', 'net_op_income_ttm', 'net_op_income_q']}
        result['is_bank'] = False
    
    try:
        info = yf.Ticker(symbol + ".IS").info
        result['industry'] = info.get('industry', '')
    except Exception:
        result['industry'] = ''
        
    return {'row_idx': row_idx, 'symbol': symbol, 'data': result}

def task_yabanci_oran(row_idx, symbol, date_a, date_b):
    return {'row_idx': row_idx, 'symbol': symbol, 'data': fetch_yabanci_oran(symbol, date_a, date_b)}

def task_historical_q(row_idx, symbol, quarter_list, live_price=None):
    """Her çeyrek için PD/DD, FD/FAVÖK (banka hariç), Borç/Özsermaye (banka hariç), İhracat Oranı; quarter_list (y,q) en güncelden eskiye.

    FD/FAVÖK paydası çeyreklik FAVÖK×4 iken piyasa değeri için Yabancı Oranı sayfasındaki güncel fiyat
    kullanılır (live_price); yoksa o çeyreğin tarihsel fiyatına düşülür. PD/DD tarihsel fiyatla kalır.
    """
    cells = []
    for (y, q) in quarter_list:
        pddd = None
        fd_favok = None
        borc_ozk = None
        ihracat_orani = None

        df_hist, resolved_h, partial_h, group_h = fetch_tr_raw_data(symbol, period_val=(y, q))
        if df_hist is None:
            cells.append((pddd, fd_favok, borc_ozk, ihracat_orani))
            continue

        res = parse_tr_data(df_hist, resolved_h, partial_h, group_h)
        if not res:
            cells.append((pddd, fd_favok, borc_ozk, ihracat_orani))
            continue

        is_bank = res.get("is_bank", False)
        hist_price = get_hist_price(symbol, y, q)
        item_2oa = res.get("item_2oa")
        ozk = res.get("ozk")

        if isinstance(hist_price, (int, float)) and isinstance(item_2oa, (int, float)):
            piyasa_degeri = hist_price * item_2oa
            if isinstance(ozk, (int, float)) and ozk != 0:
                pddd = piyasa_degeri / ozk

        if not is_bank:
            net_borc = res.get("net_borc")
            favok_q = res.get("favok_q")
            px_fd = (
                live_price
                if isinstance(live_price, (int, float)) and not isinstance(live_price, bool)
                else hist_price
            )
            if (
                isinstance(px_fd, (int, float))
                and isinstance(item_2oa, (int, float))
                and isinstance(net_borc, (int, float))
                and isinstance(favok_q, (int, float))
                and favok_q != 0
            ):
                piyasa_fd = px_fd * item_2oa
                fd = piyasa_fd + net_borc
                fd_favok = fd / (favok_q * 4)

            de = res.get("debt_equity")
            if isinstance(de, (int, float)):
                borc_ozk = de

            if q == 12:
                val_io = res.get("export_ratio_ttm")
                if isinstance(val_io, (int, float)):
                    ihracat_orani = val_io

        cells.append((pddd, fd_favok, borc_ozk, ihracat_orani))

    return {"row_idx": row_idx, "symbol": symbol, "cells": cells}


# ==============================================================================
# 6. ORKESTRASYON VE EXCEL ANA MOTORU (BELLEK ODAKLI)
# ==============================================================================


def _base_period_from_resolved(resolved):
    if not resolved or resolved == "Bulunamadı":
        return ""
    s = str(resolved).strip()
    if " (" in s:
        s = s.split(" (", 1)[0].strip()
    return s


def collect_snapshot_raw_rows(sym, d, yab_end, yab_change, piyasa_degeri):
    """Güncel özet metriklerini Raw_data uzun satırlarına dönüştürür: (ticker, metrik, dönem, değer, nf).

    PD/DD (`PDDD`) buraya yazılmaz: çeyreklik geçmişte en güncel çeyrek zaten aynı dönem
    etiketiyle dolduruluyor; anlık fiyatla tekrar yazmak çift satır ve çelişen değer üretirdi.

    `BORC_OZK` ve çeyreklik `FD_FAVOK` yalnızca çeyreklik geçmiş blokta. Snapshot’ta
    `FD_FAVOK` için yalnızca TTM satırı (`favok_ttm` paydası) eklenir.
    """
    rows = []
    base = _base_period_from_resolved(d.get("resolved_period"))
    period_q = base if base else ""
    period_ttm = f"{base} (TTM)" if base else "(TTM)"

    net_borc = d.get("net_borc", "")
    favok_ttm = d.get("favok_ttm", "")
    favok_q = d.get("favok_q", "")

    fd_ttm = ""
    if isinstance(piyasa_degeri, (int, float)) and isinstance(net_borc, (int, float)):
        fd = piyasa_degeri + net_borc
        if isinstance(favok_ttm, (int, float)) and favok_ttm != 0:
            fd_ttm = fd / favok_ttm

    def snap(code, v_q, v_t, nf):
        rows.append((sym, code, period_q, v_q, nf))
        rows.append((sym, code, period_ttm, v_t, nf))

    def snap_same(code, val, nf):
        rows.append((sym, code, period_q, val, nf))
        rows.append((sym, code, period_ttm, val, nf))

    snap("EBIT", d.get("ebit_q", ""), d.get("ebit_ttm", ""), "#,##0")

    snap("NET_OP_INC", d.get("net_op_income_q", ""), d.get("net_op_income_ttm", ""), "#,##0")

    snap("NCF", d.get("op_cash_flow_q", ""), d.get("op_cash_flow_ttm", ""), "#,##0")

    nfx = d.get("net_fx", "")
    nf_nfx = "#,##0" if isinstance(nfx, (int, float)) else None
    snap_same("NET_FX", nfx, nf_nfx)

    snap(
        "EXP_RATIO",
        d.get("export_ratio_q", ""),
        d.get("export_ratio_ttm", ""),
        "0.00%",
    )

    nf_y = "0.00%" if isinstance(yab_end, (int, float)) else None
    snap_same("YAB_END", yab_end, nf_y)
    nf_yc = "0.00%" if isinstance(yab_change, (int, float)) else None
    snap_same("YAB_CHG", yab_change, nf_yc)

    nf_pd = "#,##0" if isinstance(piyasa_degeri, (int, float)) else None
    snap_same("PD", piyasa_degeri, nf_pd)

    snap("FAVOK", favok_q, favok_ttm, "#,##0")

    # Çeyreklik FD/FAVÖK çeyreklik geçmişte (FD_FAVOK); burada yalnızca TTM paydası (favok_ttm) ile tek satır.
    nf_fd = "0.00"
    if isinstance(fd_ttm, (int, float)) and not isinstance(fd_ttm, bool):
        rows.append((sym, "FD_FAVOK", period_ttm, fd_ttm, nf_fd))

    return rows

def main_automation():
    file_name = "analysis.xlsx"
    sheet_yab_name = "Yabancı Oranı"
    print("⏳ Program başlatıldı. Excel dosyası okunuyor...")
    if not os.path.exists(file_name):
        print(f"Hata: '{file_name}' dosyası bulunamadı!")
        return
    wb = load_workbook(file_name)
    print("✅ Excel dosyası başarıyla okundu. Eski veriler temizleniyor...")

    excel_sheets.remove_facts_sheet_if_present(wb)
    excel_sheets.clear_managed_output_sheets_body(wb)

    # Çekilecek çeyrek sayısı yalnızca Config!B1 (analiz_donemi) ile belirlenir.
    window_cfg, _, _ = excel_sheets.read_config_signal_params(wb)
    num_quarters = max(1, int(window_cfg))

    main_sheet = "Main"
    if main_sheet not in wb.sheetnames:
        print(f"Hata: '{main_sheet}' sayfası bulunamadı! Tickler yalnızca bu sayfada okunur.")
        return

    ws_main = wb[main_sheet]
    tr_tasks = []
    for row in range(2, ws_main.max_row + 1):
        symbol = str(ws_main.cell(row=row, column=1).value or "").strip()
        if symbol:
            tr_tasks.append((row, symbol))

    date_a, date_b = "01-01-2024", "01-01-2024"
    if sheet_yab_name not in wb.sheetnames:
        wb.create_sheet(sheet_yab_name)
    if tr_tasks:
        test_symbol = tr_tasks[0][1]
        print(f"Yabancı oranı için geçerli iş günleri aranıyor (Referans: {test_symbol})...")
        date_a, date_b = get_yabanci_oran_dates(test_symbol)
    else:
        date_a, date_b = "01-01-2024", "01-01-2024"

    quarter_list = resolve_quarter_list_for_workbook(tr_tasks, num_quarters) if tr_tasks else []
    period_labels = [period_label(y, q) for y, q in quarter_list]
    matrix_tasks = [(r, sym, quarter_list) for r, sym in tr_tasks]

    print(
        f"\n🚀 Belleğe Alınıyor... TR: {len(tr_tasks)} şirket "
        f"(matris: {num_quarters} çeyrek, Config analiz_donemi B1)"
    )

    # ==============================================================
    # Önce TR + Yabancı (güncel fiyat), sonra çeyreklik geçmiş (FD/FAVÖK için canlı fiyat)
    # ==============================================================

    _t_start = time.perf_counter()

    tr_fin_results_map, tr_yab_results_map, hist_q_results_map = {}, {}, {}
    sym_live_price = {}

    executor_tr = concurrent.futures.ThreadPoolExecutor(max_workers=50, thread_name_prefix="TR-Mali")
    executor_yab = concurrent.futures.ThreadPoolExecutor(max_workers=50, thread_name_prefix="Yabanci")

    futures_tr = {executor_tr.submit(task_tr_stock, r, sym): ("TR", sym) for r, sym in tr_tasks}
    futures_yab = {executor_yab.submit(task_yabanci_oran, r, sym, date_a, date_b): ("YAB", sym) for r, sym in tr_tasks}
    phase1 = {}
    phase1.update(futures_tr)
    phase1.update(futures_yab)

    total_tr, total_yab = len(futures_tr), len(futures_yab)
    total_hist = len(matrix_tasks)
    done_tr, done_yab, done_hist = 0, 0, 0
    total_all = total_tr + total_yab + total_hist
    done_all = 0
    last_sym = ""

    def _bar(done, total, width=15):
        filled = int(width * done / total) if total > 0 else width
        return f"[{'█' * filled}{'░' * (width - filled)}]"

    def _print_progress():
        elapsed = time.perf_counter() - _t_start
        pct = (done_all / total_all * 100) if total_all > 0 else 100
        line = (
            f"\r  ⏳ {pct:5.1f}% {_bar(done_all, total_all, 20)} "
            f"│ TR Mali: {done_tr}/{total_tr} "
            f"│ Yabancı: {done_yab}/{total_yab} "
            f"│ Hist-Q: {done_hist}/{total_hist} "
            f"│ ⏱ {elapsed:.1f}s "
            f"│ Son: {last_sym:<10}"
        )
        sys.stdout.write(line)
        sys.stdout.flush()

    _print_progress()

    for future in concurrent.futures.as_completed(phase1):
        task_type, sym = phase1[future]
        result = future.result()
        last_sym = sym

        if task_type == "TR":
            tr_fin_results_map[sym] = result
            done_tr += 1
        else:
            tr_yab_results_map[sym] = result
            p = (result.get("data") or {}).get("price")
            if isinstance(p, (int, float)) and not isinstance(p, bool):
                sym_live_price[sym] = float(p)
            done_yab += 1

        done_all += 1
        _print_progress()

    executor_tr.shutdown(wait=False)
    executor_yab.shutdown(wait=False)

    executor_hist = concurrent.futures.ThreadPoolExecutor(max_workers=10, thread_name_prefix="Hist-Q")
    futures_hist = {
        executor_hist.submit(
            task_historical_q, r, sym, ql, sym_live_price.get(sym)
        ): ("HIST_Q", sym)
        for r, sym, ql in matrix_tasks
    }

    for future in concurrent.futures.as_completed(futures_hist):
        task_type, sym = futures_hist[future]
        hist_q_results_map[sym] = future.result()
        last_sym = sym
        done_hist += 1
        done_all += 1
        _print_progress()

    executor_hist.shutdown(wait=False)

    tr_fin_results = [tr_fin_results_map[sym] for _, sym in tr_tasks]
    tr_yab_results = [tr_yab_results_map[sym] for _, sym in tr_tasks]

    elapsed_total = time.perf_counter() - _t_start
    print(f"\n\n✅ Tüm veriler RAM'e çekildi ({elapsed_total:.1f}s). Excel'e formatlanarak yazılıyor...\n")

    yab_dict = {res["symbol"]: res["data"] for res in tr_yab_results}

    snapshot_rows_all = []
    for res in tr_fin_results:
        sym, d = res["symbol"], res["data"]
        yab_end = yab_dict[sym]["end"] if sym in yab_dict else ""
        yab_change = yab_dict[sym]["change"] if sym in yab_dict else ""
        price = yab_dict[sym]["price"] if sym in yab_dict else ""
        item_2oa = d.get("item_2oa", "")

        piyasa_degeri = ""
        if isinstance(price, (int, float)) and isinstance(item_2oa, (int, float)):
            piyasa_degeri = price * item_2oa

        snapshot_rows_all.extend(
            collect_snapshot_raw_rows(sym, d, yab_end, yab_change, piyasa_degeri)
        )

    # Yabancı Oranı Sayfasını Yazma
    if sheet_yab_name in wb.sheetnames:
        ws_yab = wb[sheet_yab_name]
        headers_yab = {1: "Ticker", 2: "Fiyat", 3: date_a, 4: date_b, 5: "Change", 6: "Effect"}
        for col_num, text in headers_yab.items(): ws_yab.cell(row=1, column=col_num, value=text)
        
        for idx, res in enumerate(tr_yab_results, start=2):
            sym, yd = res['symbol'], res['data']
            ws_yab.cell(row=idx, column=1, value=sym)
            c_price = ws_yab.cell(row=idx, column=2, value=yd['price'])
            c_start = ws_yab.cell(row=idx, column=3, value=yd['start'])
            c_end = ws_yab.cell(row=idx, column=4, value=yd['end'])
            c_change = ws_yab.cell(row=idx, column=5, value=yd['change'])
            c_effect = ws_yab.cell(row=idx, column=6, value=yd['effect'])
            
            if isinstance(c_price.value, (int, float)): c_price.number_format = '#,##0.00'
            for c in [c_start, c_end, c_change, c_effect]:
                if isinstance(c.value, (int, float)): c.number_format = '0.00%'

    nq = len(quarter_list)
    matrix_rows = []
    for r, sym in tr_tasks:
        pack = hist_q_results_map.get(sym, {})
        cells = pack.get("cells", [])
        if len(cells) < nq:
            cells = cells + [(None, None, None, None)] * (nq - len(cells))
        else:
            cells = cells[:nq]
        matrix_rows.append((r, sym, cells))

    excel_sheets.ensure_config_sheet(wb)
    excel_sheets.write_raw_data(wb, period_labels, matrix_rows, snapshot_rows_all)
    excel_sheets.write_sinyal_sheet(
        wb, tr_tasks, tr_fin_results_map, matrix_rows, period_labels
    )

    wb.save(file_name)
    print(f"✅ Excel dosyası başarıyla güncellendi ve kapatıldı!")


if __name__ == "__main__":
    main_automation()